from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_practice_excel(form_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "F2-INICIACIÓN"

    # Estilos
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_aligned = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    # ----------------- ENCABEZADO -----------------
    ws.merge_cells('D1:N1')
    ws['D1'] = "FORMATO DE INICIACIÓN Y COMPROMISO DE LA ETAPA PRACTICA"
    ws['D1'].font = title_font
    ws['D1'].alignment = center_aligned

    ws.merge_cells('D2:N2')
    ws['D2'] = "PROCESO: Prestación del Servicio / Extensión y Proyección a la Comunidad"
    ws['D2'].alignment = center_aligned

    ws['F3'] = "Código: FPESF-4.2-52"
    ws['H3'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
    ws['J3'] = "Version: 1.0"

    # ----------------- 1. DATOS DEL ESTUDIANTE -----------------
    ws['B5'] = "1. DATOS DEL ESTUDIANTE DE PRACTICA"
    ws['B5'].font = bold_font
    ws['B5'].fill = header_fill

    # Nombre
    ws['B6'] = "Nombre completo del estudiante:"
    ws['D6'] = form_data['student'].get('fullName', '')

    # Tipo documento
    ws['B7'] = "Tipo de documento:"
    cc_checked = 'X' if form_data['student'].get('documentType') == 'CC' else '_'
    ce_checked = 'X' if form_data['student'].get('documentType') == 'CE' else '_'
    ws['D7'] = f"C.C. {cc_checked}     C.E. {ce_checked}"
    ws['I7'] = "Número de documento:"
    ws['K7'] = form_data['student'].get('documentNumber', '')

    # Nivel formación
    ws['B8'] = "Nivel de formación:"
    tecnologia_checked = 'X' if form_data['student'].get('educationLevel') == 'Tecnología' else '_'
    profesional_checked = 'X' if form_data['student'].get('educationLevel') == 'Profesional Universitario' else '_'
    ws['D8'] = f"Tecnología: {tecnologia_checked}"
    ws['J8'] = f"Profesional Universitario: {profesional_checked}"

    # Programa académico
    ws['B9'] = "Programa académico:"
    ws['D9'] = form_data['student'].get('academicProgram', '')

    # Contacto
    ws['B10'] = "Teléfonos de contacto:"
    ws['D10'] = form_data['student'].get('phone', '')
    ws['H10'] = "Correo electrónico FESC:"
    ws['J10'] = form_data['student'].get('email', '')

    # ----------------- 2. DATOS DE LA EMPRESA -----------------
    ws['B12'] = "2. DATOS GENERALES DE LA EMPRESA"
    ws['B12'].font = bold_font
    ws['B12'].fill = header_fill

    # Nombre empresa
    ws['B13'] = "Nombre de la Empresa:"
    ws['D13'] = form_data['company'].get('name', '')

    # Representante legal
    ws['B14'] = "Representante legal:"
    ws['D14'] = form_data['company'].get('legalRepresentative', '')

    # NIT
    ws['B15'] = "NIT:"
    ws['D15'] = form_data['company'].get('nit', '')

    # Actividad económica
    ws['B16'] = "Actividad económica:"
    ws['D16'] = form_data['company'].get('economicActivity', '')

    # Sector económico
    ws['B17'] = "Sector Económico:"
    ws['D17'] = form_data['company'].get('economicSector', '')

    # Nacionalidad
    ws['B18'] = "Nacionalidad de la empresa:"
    ws['D18'] = form_data['company'].get('nationality', '')

    # Tamaño empresa
    ws['B19'] = "Tamaño de la Empresa:"
    ws['D19'] = form_data['company'].get('size', '')

    # Página web
    ws['B20'] = "Pagina Web:"
    ws['D20'] = form_data['company'].get('website', '')

    # Dirección
    ws['B21'] = "Dirección:"
    ws['D21'] = form_data['company'].get('address', '')

    # Teléfono
    ws['B22'] = "Teléfonos de contacto:"
    ws['D22'] = form_data['company'].get('phone', '')

    # ----------------- 3. JEFE INMEDIATO -----------------
    ws['B24'] = "3. INFORMACIÓN DEL JEFE INMEDIATO DEL PRACTICANTE EN LA EMPRESA:"
    ws['B24'].font = bold_font
    ws['B24'].fill = header_fill

    # Nombre jefe
    ws['B25'] = "Nombre completo:"
    ws['D25'] = form_data['boss'].get('fullName', '')

    # Cargo
    ws['B26'] = "Cargo:"
    ws['D26'] = form_data['boss'].get('position', '')

    # Email
    ws['B27'] = "Correo electrónico:"
    ws['D27'] = form_data['boss'].get('email', '')

    # Teléfono
    ws['B28'] = "Teléfonos de contacto"
    ws['D28'] = form_data['boss'].get('phone', '')

    # ----------------- 4. ESCENARIO DE PRÁCTICA -----------------
    ws['B30'] = "4. ESCENARIO DE PRÁCTICA"
    ws['B30'].font = bold_font
    ws['B30'].fill = header_fill

    # Modalidad
    ws['B31'] = "Por favor seleccione la modalidad de la práctica que va a desarrollar (Solo una opción)"
    ws['B32'] = "1. MODALIDAD CONVENIO"
    ws['F32'] = "2. MODALIDAD VINCULACIÓN LABORAL"
    ws['J32'] = "3. MODALIDAD INVESTIGACIÓN"
    ws['N32'] = "4. MODALIDAD PROYECTO PRODUCTIVO O SOCIAL"

    # Submodalidades
    convenio_checked = 'X' if form_data['practice'].get('modality') == 'CONVENIO' else '_'
    vinculacion_checked = 'X' if form_data['practice'].get('modality') == 'VINCULACION' else '_'
    investigacion_checked = 'X' if form_data['practice'].get('modality') == 'INVESTIGACION' else '_'
    proyecto_checked = 'X' if form_data['practice'].get('modality') == 'PROYECTO' else '_'

    ws['B33'] = f"Práctica (Regional o Nacional) {convenio_checked}"
    ws['D33'] = f"Práctica Internacional {convenio_checked}"
    ws['F33'] = f"Práctica contrato de aprendizaje {vinculacion_checked}"
    ws['H33'] = f"Propuesta mejora empresarial {investigacion_checked}"
    ws['J33'] = f"Auxiliar de investigacion {investigacion_checked}"
    ws['L33'] = f"Plan de mejora para su emprendimiento {proyecto_checked}"
    ws['N33'] = f"Plan de Negocio {proyecto_checked}"

    # Cargo
    ws['B34'] = "Cargo del practicante:"
    ws['D34'] = form_data['practice'].get('position', '')

    # Departamento
    ws['B35'] = "Área o Departamento:"
    ws['D35'] = form_data['practice'].get('department', '')

    # Periodo
    ws['B36'] = "Periodo de la práctica:"
    ws['D36'] = f"Fecha de inicio: {form_data['practice'].get('startDate', '')}        Fecha fin: {form_data['practice'].get('endDate', '')}"
    ws['D37'] = "En el nivel de tecnología el estudiante desarrollará un mínimo de 480 horas y en el nivel profesional desarrollará un mínimo de 640 horas"

    # Horario
    ws['B38'] = "Horario de labores concertado:"
    ws['D38'] = form_data['practice'].get('workSchedule', '')

    # Remuneración
    ws['B39'] = "Práctica remunerada:"
    remunerada_checked = 'X' if form_data['practice'].get('isPaid') else '_'
    no_remunerada_checked = 'X' if not form_data['practice'].get('isPaid') else '_'
    ws['D39'] = f"SI: {remunerada_checked}"
    ws['G39'] = f"NO: {no_remunerada_checked}"
    ws['I39'] = f"Valor mensual de la remuneración: {form_data['practice'].get('salary', '')}"

    # Uniforme
    ws['B40'] = "Suministra uniforme de dotacion al practicante:"
    uniforme_si_checked = 'X' if form_data['practice'].get('providesUniform') else '_'
    uniforme_no_checked = 'X' if not form_data['practice'].get('providesUniform') else '_'
    ws['D40'] = f"SI: {uniforme_si_checked}"
    ws['G40'] = f"NO: {uniforme_no_checked}"

    # Funciones
    ws['B41'] = "Funciones del practicante (Aplicable a todas las modalidades de práctica)"
    ws['D41'] = form_data['practice'].get('functions', '')

    # Recursos
    ws['B42'] = "Recursos a disposición del practicante:"
    ws['D42'] = "Computador y equipo de oficina" if form_data['practice'].get('resources', {}).get('computer') else ''
    ws['G42'] = "Otros"
    ws['I42'] = form_data['practice'].get('resources', {}).get('othersDescription', '') if form_data['practice'].get('resources', {}).get('others') else ''

    # Asesor
    ws['B43'] = "Docente asesor FESC asignado:"
    ws['D43'] = form_data['practice'].get('advisor', '')

    # ----------------- 5. COMPROMISOS -----------------
    ws['B45'] = "5. COMPROMISOS"
    ws['B45'].font = bold_font
    ws['B45'].fill = header_fill

    compromisos_texto = [
        "Con la firma de este documento, me comprometo hasta el día de la culminación de la Práctica cumplir respetuosa y responsablemente con:",
        "1. Los informes de practica en cada corte solicitados por el Asesor de Practica oportunamente según las fechas establecidas en el calendario académico y con las calidades esperadas segun mi plan de trabajo de practica.",
        "2. (Solo aplica para estudiantes por convenio) Las funciones asignadas, acatar las políticas internas de la empresa y del Reglamento de Prácticas FESC. De igual manera informaré oportunamente por escrito al jefe Inmediato y a la Coordinación de Prácticas cualquier solicitud o permiso de inasistencia, ya sea por enfermedad, calamidad familiar o actividad académica, presentando el respectivo soporte."
    ]

    for i, texto in enumerate(compromisos_texto):
        ws[f'B{46 + i}'] = texto

    # ----------------- 6. FIRMAS -----------------
    ws['B50'] = "6. FIRMAS"
    ws['B50'].font = bold_font
    ws['B50'].fill = header_fill

    ws['B52'] = "Estudiante Practicante"
    ws['F52'] = "Jefe Inmediato del practicante"
    ws['J52'] = "Docente Asesor Asignado"

    # Aplicar bordes y estilos a todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical='center')

    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    return wb